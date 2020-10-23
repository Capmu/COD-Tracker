'''
    project   : CPMU-X001 : COD-Tracker
    version   : 1.0
    developer : Capmu

    Unique Statement
        - line(...............) : Convert float to date.
'''

#--------------------------------------------------------
# Classes / Libraries / Packages
#--------------------------------------------------------
from Classes.DeliveryInfo import DeliveryInfo

import os
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from shutil import copyfile, move
import datetime
import time

#--------------------------------------------------------
# Fucntions
#--------------------------------------------------------
def logInfo(situation):
    if situation == "intro":
        print("")
    if situation == "outro":
        print("-----------------------------------------")
        print("                 Success                 ")
        print("-----------------------------------------")

        time.sleep(300)

    return()

def Number_of_cell_alphabet(alphabetNumber):

    if alphabetNumber == 1:
        thisAlphabet = 'A'
    elif alphabetNumber == 2:
        thisAlphabet = 'B'
    elif alphabetNumber == 3:
        thisAlphabet = 'C'
    elif alphabetNumber == 4:
        thisAlphabet = 'D'
    elif alphabetNumber == 5:
        thisAlphabet = 'E'
    elif alphabetNumber == 6:
        thisAlphabet = 'F'
    elif alphabetNumber == 7:
        thisAlphabet = 'G'
    elif alphabetNumber == 8:
        thisAlphabet = 'H'
    elif alphabetNumber == 9:
        thisAlphabet = 'I'
    elif alphabetNumber == 10:
        thisAlphabet = 'J'
    elif alphabetNumber == 11:
        thisAlphabet = 'K'
    elif alphabetNumber == 12:
        thisAlphabet = 'L'
    elif alphabetNumber == 13:
        thisAlphabet = 'M'
    elif alphabetNumber == 14:
        thisAlphabet = 'N'
    elif alphabetNumber == 15:
        thisAlphabet = 'O'
    elif alphabetNumber == 16:
        thisAlphabet = 'P'
    elif alphabetNumber == 17:
        thisAlphabet = 'Q'
    elif alphabetNumber == 18:
        thisAlphabet = 'R'
    elif alphabetNumber == 19:
        thisAlphabet = 'S'
    elif alphabetNumber == 20:
        thisAlphabet = 'T'
    elif alphabetNumber == 21:
        thisAlphabet = 'U'
    elif alphabetNumber == 22:
        thisAlphabet = 'V'
    elif alphabetNumber == 23:
        thisAlphabet = 'W'
    elif alphabetNumber == 24:
        thisAlphabet = 'X'
    elif alphabetNumber == 25:
        thisAlphabet = 'Y'
    elif alphabetNumber == 26:
        thisAlphabet = 'Z'
    else:
        print("Uncorrect alphabet number !")
    
    if thisAlphabet:
        return(thisAlphabet)

def displayInvalidPayment(invalidPayments):

    print("-----------------------------------------")
    print("             Invalid Payment             ")
    print("-----------------------------------------")

    for showInfo in invalidPayments:
        print(showInfo.sendingDate, "   ", showInfo.deliveryCode)
    
    print("-----------------------------------------\n")

    return()

def getAllFileNameAtPath(path):
    for filePath in os.walk(path): #full example --> for root, dirs, files in os.walk("./Checking File"):
        for files in filePath:
            pass

    return(files)

def readExcelAtSheet(path, sheetNumber):
    return(xlrd.open_workbook(path).sheet_by_index(sheetNumber)) #for [xlrd] library

def getSendingInfo(path, bias):

    excelFiles = getAllFileNameAtPath(path)
    sendings = []
    
    print(" -> Loaded file.")
    for excelFile in excelFiles:
        
        reader = readExcelAtSheet(path + "/" + excelFile, 0)
        
        for i in range(len(reader.col_values(0)) - 1 + bias):
            #pull informations from excel to python list
            sending = DeliveryInfo()

            #=======================================================================================================================================
            # Unique Statement
            #=======================================================================================================================================
            sending.sendingDate = datetime.datetime.utcfromtimestamp(((reader.col_values(columnNumberSendingDate)[i + 1]) - 25569) * 86400.0).date()
            #=======================================================================================================================================
            sending.deliveryCode = str(reader.col_values(columnNumberDeliveryCode_sending)[i + 1])
            sending.expectedCOD = reader.col_values(columnNumberExpectedCOD)[i + 1]

            if len(listOfAdditionalColumnName) > 0:

                sending.setCustomerInfoDict(additionalInfoValueDict.copy())

                for columnName in listOfAdditionalColumnName:
                    sending.customerInfoDict[columnName] = reader.col_values(additionalInfoColumnNumberDict[columnName])[i + 1]
            
            sendings.append(sending)
    
        print("     ", str(excelFile))
        
    print("")

    return(sendings)

def getReceivingInfo():

    print(" -> Loaded receiving file.")

    excelFiles = getAllFileNameAtPath(excelProductReceivedPath)
    receivingDict = {}

    for excelFile in excelFiles:

        reader = readExcelAtSheet(excelProductReceivedPath + "/" + excelFile, 0)

        for i in range(len(reader.col_values(0)) - 1 + columnBias):
            #pull using informations from excel to python list (Only DeliveryCode and it's COD)
            receivingDict[str(reader.col_values(columnNumberDeliveryCode_receiving)[i + 1])] = reader.col_values(columnNumberActualCOD)[i + 1]
        
        print("     ", str(excelFile))
        
    print("")

    return(receivingDict)

def getColumnNames(filePath, sheetNumber):
    
    reader = readExcelAtSheet(filePath, 0)
    columnNames = []

    for i in range(len(reader.row_values(0))):
        columnNames.append(str(reader.col_values(i)))
    
    return columnNames

def initialReportFile():

    copyfile(reportTemplatePath, excelReportPath)

    return()

def appendSendingInfo(path, sendings):
    
    reader = readExcelAtSheet(path, 0)
    reportWorkbook = load_workbook(path)      #for [openpyxl] libraly
    reportRecorder = reportWorkbook.active

    startRow = len(reader.col_values(0)) + 1 #for "OpenPyXl" format

    for i in range(len(sendings)):
        reportRecorder[Number_of_cell_alphabet(columnNumberSendingDate + 1) + str(startRow + i)] = sendings[i].sendingDate
        reportRecorder[Number_of_cell_alphabet(columnNumberDeliveryCode_sending + 1) + str(startRow + i)] = sendings[i].deliveryCode
        reportRecorder[Number_of_cell_alphabet(columnNumberExpectedCOD + 1) + str(startRow + i)] = sendings[i].expectedCOD
        
        if len(listOfAdditionalColumnName) > 0:

            for columnName in listOfAdditionalColumnName:
                reportRecorder[Number_of_cell_alphabet(additionalInfoColumnNumberDict[columnName] + 1) + str(startRow + i)] = sendings[i].customerInfoDict[columnName]

    reportWorkbook.save(path)

    return()

def updateReport():

    readyMessage = " -> Ready to check.\n"

    #check if need to create the report file.
    if os.path.exists(excelReportPath):
        print(" -> Report file is already created.\n")
        print(readyMessage)
    else:
        initialReportFile()
        print(" -> Created report file.\n")
        print(readyMessage)

    appendSendingInfo(excelReportPath, sendings)
    
    return()

def getPaymentList(paymentReceivingDict):
    
    paidList = []
    nonPaidList = []
    remainingReceivingDict = []
    invalidPaidList = []

    print(len(paymentReceivingDict))

    for sending in sendingDatabase:

        if sending.deliveryCode in paymentReceivingDict:

            sending.actualCOD = paymentReceivingDict[sending.deliveryCode]
            sending.checkStatus()

            if sending.status == "success-payment":
                paidList.append(sending)

            elif sending.status == "invalid-COD":
                invalidPaidList.append(sending)
            
            paymentReceivingDict.pop(sending.deliveryCode)
        
        else:
            nonPaidList.append(sending)

    remainingReceivingDict = paymentReceivingDict

    #inform invalid payment
    if len(invalidPaidList) > 0:
        displayInvalidPayment(invalidPaidList)

    return(paidList, nonPaidList, remainingReceivingDict)

def trackCOD():

    reportWorkbook = load_workbook(excelReportPath)      #for [openpyxl] libraly
    reportRecorder = reportWorkbook.active

    lenOfHighLight = len(getColumnNames(excelReportPath, 0))

    for i in range(len(sendingDatabase)):

        if sendingDatabase[i].status == "success-payment":
            for columnNumber in range(lenOfHighLight):
                reportRecorder[Number_of_cell_alphabet(columnNumber + 1) + str(i + 2)].fill = lightGreen_fill
        
        elif sendingDatabase[i].status == "invalid-COD":
            for columnNumber in range(lenOfHighLight):
                reportRecorder[Number_of_cell_alphabet(columnNumber + 1) + str(i + 2)].fill = lightRed_fill

    reportWorkbook.save(excelReportPath)

    print(" -> marked.\n")

    return()

def createRemainingReceivingFile():
    
    copyfile(reportTemplatePath, remainReceivingFilePath)

    reportWorkbook = load_workbook(remainReceivingFilePath)      #for [openpyxl] libraly
    reportRecorder = reportWorkbook.active

    for i in range(len(remainingReceivingDict)):
        reportRecorder[Number_of_cell_alphabet(columnNumberDeliveryCode_sending + 1) + str(i + 2)] = list(remainingReceivingDict)[i]
        reportRecorder[Number_of_cell_alphabet(columnNumberExpectedCOD + 1) + str(i + 2)] = remainingReceivingDict[list(remainingReceivingDict)[i]]
    #===================================================================================================================================
    # Unique Statement
    #===================================================================================================================================
    reportRecorder[Number_of_cell_alphabet(4) + str(len(remainingReceivingDict) + 2)] = "COD-Tracker"
    #===================================================================================================================================

    reportWorkbook.save(remainReceivingFilePath)

    return()

def moveUsedFiles():

    sendingFiles = getAllFileNameAtPath(excelProductSendingPath)
    receivingFiles = getAllFileNameAtPath(excelProductReceivedPath)
    
    for sendingFile in sendingFiles:
        move(excelProductSendingPath + "/" + sendingFile, excelBackup_ProductSendingPath + "/" + sendingFile)
    
    for receivingFile in receivingFiles:
        move(excelProductReceivedPath + "/" + receivingFile, excelBackup_ProductReceivedPath + "/" + receivingFile)

    if len(remainingReceivingDict) > 0:
        createRemainingReceivingFile()

    print(" -> moved file.\n")

    return()

#--------------------------------------------------------
# Variables
#--------------------------------------------------------
reportName = "เช็คยอด-COD.xlsx"
summaryReportName = "เช็คยอด-COD-(สรุป).xlsx"
reportFolder = "3. ไฟล์เช็คยอด/"

excelProductSendingPath = "1. ไฟล์วันที่-ส่งของ"
excelProductReceivedPath = "2. ไฟล์วันที่-ลูกค้ารับของ"
excelReportPath = reportFolder + reportName
excelSummaryReportPath = reportFolder + "สรุป/" + summaryReportName

excelBackup_ProductSendingPath = "source-code/ประวัติการดำเนินการ/1. ไฟล์วันที่-ส่งของ (Backup)"
excelBackup_ProductReceivedPath = "source-code/ประวัติการดำเนินการ/2. ไฟล์วันที่-ลูกค้ารับของ (Backup)"
reportTemplatePath = "source-code/python-code/support-files/เช็คยอด-COD.xlsx"

remainReceivingFilePath = excelProductReceivedPath + "/คงเหลือ.xlsx"

lightGreen_fill = PatternFill(start_color='d3ffd1', end_color='d3ffd1', fill_type='solid')
lightRed_fill = PatternFill(start_color='ffd6d6', end_color='ffd6d6', fill_type='solid')

#========================================================
# Configuration
#========================================================

# Column Settings for sending files.
columnNumberSendingDate = 0
columnNumberDeliveryCode_sending = 1
columnNumberExpectedCOD = 3

# Column Settings for receiving files.
columnNumberDeliveryCode_receiving = 1
columnNumberActualCOD = 3

# Addition Infomation : up to each project.

listOfAdditionalColumnName = ["customerName"]

additionalInfoColumnNumberDict = {
    listOfAdditionalColumnName[0] : 2
}
additionalInfoValueDict = {
    listOfAdditionalColumnName[0] : "default-value"
}

columnBias = -1 #from summary cash.
#--------------------------------------------------------
# Implementation
#--------------------------------------------------------
logInfo("intro")
sendings = getSendingInfo(excelProductSendingPath, columnBias)
receivingDict = getReceivingInfo()
updateReport()
sendingDatabase = getSendingInfo(reportFolder, 0) #re-use this function, so have some wired structure.
paidList, nonPaidList, remainingReceivingDict = getPaymentList(receivingDict)
trackCOD()
moveUsedFiles()
logInfo("outro")
