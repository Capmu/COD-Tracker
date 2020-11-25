'''
    project   : CPMU-X001 : COD-Tracker
    version   : 2020.10.1
    developer : Capmu

    Unique Statement
        - line(139) : Convert float to date.
        - line(307) : Add footer.
'''

#--------------------------------------------------------
# Classes / Libraries / Packages
#--------------------------------------------------------
from Classes.DeliveryInfo import DeliveryInfo

import os
import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from shutil import copyfile, move
import datetime
import time

#--------------------------------------------------------
# Fucntions
#--------------------------------------------------------
def logInfo(situation):
    if situation == "intro":
        print("")
    if situation == "outro":
        print("-----------------------------------------")
        print("                 Success                 ")
        print("-----------------------------------------")

        time.sleep(300)

    return()

def Number_of_cell_alphabet(alphabetNumber):

    if alphabetNumber == 1:
        thisAlphabet = 'A'
    elif alphabetNumber == 2:
        thisAlphabet = 'B'
    elif alphabetNumber == 3:
        thisAlphabet = 'C'
    elif alphabetNumber == 4:
        thisAlphabet = 'D'
    elif alphabetNumber == 5:
        thisAlphabet = 'E'
    elif alphabetNumber == 6:
        thisAlphabet = 'F'
    elif alphabetNumber == 7:
        thisAlphabet = 'G'
    elif alphabetNumber == 8:
        thisAlphabet = 'H'
    elif alphabetNumber == 9:
        thisAlphabet = 'I'
    elif alphabetNumber == 10:
        thisAlphabet = 'J'
    elif alphabetNumber == 11:
        thisAlphabet = 'K'
    elif alphabetNumber == 12:
        thisAlphabet = 'L'
    elif alphabetNumber == 13:
        thisAlphabet = 'M'
    elif alphabetNumber == 14:
        thisAlphabet = 'N'
    elif alphabetNumber == 15:
        thisAlphabet = 'O'
    elif alphabetNumber == 16:
        thisAlphabet = 'P'
    elif alphabetNumber == 17:
        thisAlphabet = 'Q'
    elif alphabetNumber == 18:
        thisAlphabet = 'R'
    elif alphabetNumber == 19:
        thisAlphabet = 'S'
    elif alphabetNumber == 20:
        thisAlphabet = 'T'
    elif alphabetNumber == 21:
        thisAlphabet = 'U'
    elif alphabetNumber == 22:
        thisAlphabet = 'V'
    elif alphabetNumber == 23:
        thisAlphabet = 'W'
    elif alphabetNumber == 24:
        thisAlphabet = 'X'
    elif alphabetNumber == 25:
        thisAlphabet = 'Y'
    elif alphabetNumber == 26:
        thisAlphabet = 'Z'
    else:
        print("Uncorrect alphabet number !")
    
    if thisAlphabet:
        return(thisAlphabet)

def displayInvalidPayment(invalidPayments):

    print("-----------------------------------------")
    print("             Invalid Payment             ")
    print("-----------------------------------------")

    for showInfo in invalidPayments:
        print(showInfo.sendingDate, "   ", showInfo.deliveryCode, "   |   Expected: ", showInfo.expectedCOD, "   ", "Actual: ", showInfo.actualCOD)
    
    print("-----------------------------------------\n")

    return()

def getAllFileNameAtPath(path):
    for filePath in os.walk(path): #full example --> for root, dirs, files in os.walk("./Checking File"):
        for files in filePath:
            pass

    return(files)

def readExcelAtSheet(path, sheetNumber):
    return(xlrd.open_workbook(path).sheet_by_index(sheetNumber)) #for [xlrd] library

def getSendingInfo(path, bias):

    excelFiles = getAllFileNameAtPath(path)
    sendings = []
    
    for excelFile in excelFiles:

        biasForThisFile = bias
        
        reader = readExcelAtSheet(path + "/" + excelFile, 0)
        
        for i in range(len(reader.col_values(0)) - 1 + biasForThisFile):

            #pull informations from excel to python list
            sending = DeliveryInfo()
            
            #============================================================================================
            # Unique Statement | bias = 0 means use Old format excel column topic. must be manual adjust
            #============================================================================================
            if biasForThisFile == -1:
                sending.sendingDate = str(reader.col_values(columnNumberSendingDate)[i + 1])
                sending.deliveryCode = str(reader.col_values(columnNumberDeliveryCode_sending)[i + 1])
                sending.expectedCOD = reader.col_values(columnNumberExpectedCOD)[i + 1]
            elif biasForThisFile == 0:
                sending.sendingDate = str(reader.col_values(0)[i + 1])
                sending.deliveryCode = str(reader.col_values(1)[i + 1])
                sending.expectedCOD = reader.col_values(3)[i + 1]
            #============================================================================================

            if len(listOfAdditionalColumnName) > 0:

                sending.setCustomerInfoDict(additionalInfoValueDict.copy())

                for columnName in listOfAdditionalColumnName:
                    #===============================================================================================================
                    # Unique Statement | bias = 0 means use Old format excel column topic. must be manual adjust
                    #===============================================================================================================
                    if biasForThisFile == -1:
                        sending.customerInfoDict[columnName] = reader.col_values(additionalInfoColumnNumberDict[columnName])[i + 1]
                    elif biasForThisFile == 0:
                        sending.customerInfoDict[columnName] = reader.col_values(2)[i + 1]
                    #===============================================================================================================

            sendings.append(sending)

    return(sendings)

def getReceivingInfo():

    excelFiles = getAllFileNameAtPath(excelProductReceivedPath)
    receivingDict = {}

    for excelFile in excelFiles:

        reader = readExcelAtSheet(excelProductReceivedPath + "/" + excelFile, 0)

        for i in range(len(reader.col_values(0)) - 1 + receivingFooterBias - receivingHeaderBias):
            #pull using informations from excel to python list (Only DeliveryCode and it's COD)
            receivingDict[str(reader.col_values(columnNumberDeliveryCode_receiving)[i + 1 + receivingHeaderBias])] = reader.col_values(columnNumberActualCOD)[i + 1 + receivingHeaderBias]

    return(receivingDict)

def getColumnNames(filePath, sheetNumber):
    
    reader = readExcelAtSheet(filePath, 0)
    columnNames = []

    for i in range(len(reader.row_values(0))):
        columnNames.append(str(reader.col_values(i)))
    
    return columnNames

def initialReportFile():

    copyfile(reportTemplatePath, excelReportPath)

    return()

def appendSendingInfo(path, sendings, sheetName):

    reader = readExcelAtSheet(path, 0)
    reportWorkbook = load_workbook(path)      #for [openpyxl] libraly
    reportRecorder = reportWorkbook[sheetName]

    startRow = len(reader.col_values(0)) + 1 #for "OpenPyXl" format

    #=================================================================================================================================================================
    # Unique Statement | cuz don't use all data in excel | so just fix the column.
    #=================================================================================================================================================================
    for i in range(len(sendings)):
        reportRecorder["A" + str(startRow + i)] = sendings[i].sendingDate
        reportRecorder["B" + str(startRow + i)] = sendings[i].deliveryCode
        reportRecorder["D" + str(startRow + i)] = sendings[i].expectedCOD
        
        if len(listOfAdditionalColumnName) > 0:

            # for columnName in listOfAdditionalColumnName:
            #     reportRecorder["C" + str(startRow + i)] = sendings[i].customerInfoDict[columnName]

            for columnName in listOfAdditionalColumnName:
                if columnName == "customerName":
                    reportRecorder["C" + str(startRow + i)] = sendings[i].customerInfoDict[columnName]
                elif columnName == "phoneNumber":
                    reportRecorder["F" + str(startRow + i)] = sendings[i].customerInfoDict[columnName]
    #=================================================================================================================================================================

    reportWorkbook.save(path)

    return()

def updateReport():

    readyMessage = " -> This program is working . . .\n"

    #check if need to create the report file.
    if os.path.exists(excelReportPath):
        print(" -> Report file is already created.\n")
        print(readyMessage)
    else:
        initialReportFile()
        print(" -> Created report file.\n")
        print(readyMessage)

    appendSendingInfo(excelReportPath, sendings, databaseSheetName)
    
    return()

def getPaymentList(paymentReceivingDict):
    
    paidList = []
    nonPaidList = []
    remainingReceivingDict = []
    invalidPaidList = []

    for sending in sendingDatabase:

        if sending.deliveryCode in paymentReceivingDict:

            sending.actualCOD = paymentReceivingDict[sending.deliveryCode]
            sending.checkStatus()

            if sending.status == "success-payment":
                paidList.append(sending)

            elif sending.status == "non-COD":
                paidList.append(sending)

            elif sending.status == "invalid-COD":
                invalidPaidList.append(sending)
            
            paymentReceivingDict.pop(sending.deliveryCode)
        
        else:
            nonPaidList.append(sending)

    remainingReceivingDict = paymentReceivingDict

    #inform invalid payment
    if len(invalidPaidList) > 0:
        displayInvalidPayment(invalidPaidList)

    return(paidList, nonPaidList, remainingReceivingDict)

def trackCOD():

    reportWorkbook = load_workbook(excelReportPath)      #for [openpyxl] libraly
    reportRecorder = reportWorkbook.active

    lenOfHighLight = len(getColumnNames(excelReportPath, 0))

    for i in range(len(sendingDatabase)):

        if sendingDatabase[i].status == "success-payment":
            for columnNumber in range(lenOfHighLight):
                reportRecorder[Number_of_cell_alphabet(columnNumber + 1) + str(i + 2)].fill = lightGreen_fill
        
        elif sendingDatabase[i].status == "non-COD":
            for columnNumber in range(lenOfHighLight):
                reportRecorder[Number_of_cell_alphabet(columnNumber + 1) + str(i + 2)].fill = lightBlack_fill
        
        elif sendingDatabase[i].status == "invalid-COD":
            for columnNumber in range(lenOfHighLight):
                reportRecorder[Number_of_cell_alphabet(columnNumber + 1) + str(i + 2)].fill = lightRed_fill
        
        else:
            print(sendingDatabase[i].status)

    reportWorkbook.save(excelReportPath)

    print(" -> checked.\n")

    return()

def createRemainingReceivingFile():
    
    copyfile(remainingReportTemplatePath, remainReceivingFilePath)

    reportWorkbook = load_workbook(remainReceivingFilePath)      #for [openpyxl] libraly
    reportRecorder = reportWorkbook.active

    #========================================================================
    # Unique Statement | make structure like receiving files.
    #========================================================================
    for i in range(receivingHeaderBias):
        reportRecorder[Number_of_cell_alphabet(3) + str(i + 2)] = "."

    for i in range(len(remainingReceivingDict) - receivingFooterBias):

        if i <= len(remainingReceivingDict) - 1:
            reportRecorder[Number_of_cell_alphabet(columnNumberDeliveryCode_receiving + 1) + str(i + 2 + receivingHeaderBias)] = list(remainingReceivingDict)[i]
            reportRecorder[Number_of_cell_alphabet(columnNumberActualCOD + 1) + str(i + 2 + receivingHeaderBias)] = remainingReceivingDict[list(remainingReceivingDict)[i]]
        else:
            reportRecorder[Number_of_cell_alphabet(3) + str(i + 2 + receivingHeaderBias)] = "."
    #========================================================================

    reportWorkbook.save(remainReceivingFilePath)

    return()

def moveUsedFiles():

    sendingFiles = getAllFileNameAtPath(excelProductSendingPath)
    receivingFiles = getAllFileNameAtPath(excelProductReceivedPath)
    
    for sendingFile in sendingFiles:
        move(excelProductSendingPath + "/" + sendingFile, excelBackup_ProductSendingPath + "/" + sendingFile)
    
    for receivingFile in receivingFiles:
        move(excelProductReceivedPath + "/" + receivingFile, excelBackup_ProductReceivedPath + "/" + receivingFile)

    if len(remainingReceivingDict) > 0:
        createRemainingReceivingFile()

    return()

#--------------------------------------------------------
# Variables
#--------------------------------------------------------
reportName = "เช็คยอด-COD.xlsx"
reportFolder = "3. ไฟล์เช็คยอด/"

excelProductSendingPath = "1. ไฟล์วันที่-ส่งของ"
excelProductReceivedPath = "2. ไฟล์วันที่-ลูกค้ารับของ"
excelReportPath = reportFolder + reportName

excelBackup_ProductSendingPath = "source-code/ประวัติการดำเนินการ/1. ไฟล์วันที่-ส่งของ (Backup)"
excelBackup_ProductReceivedPath = "source-code/ประวัติการดำเนินการ/2. ไฟล์วันที่-ลูกค้ารับของ (Backup)"
reportTemplatePath = "source-code/python-code/support-files/เช็คยอด-COD.xlsx"
remainingReportTemplatePath = "source-code/python-code/support-files/เช็คยอด-COD-คงเหลือ.xlsx"

remainReceivingFilePath = excelProductReceivedPath + "/คงเหลือ.xlsx"

databaseSheetName = "รวม"

lightGreen_fill = PatternFill(start_color='d3ffd1', end_color='d3ffd1', fill_type='solid')
lightRed_fill = PatternFill(start_color='ffd6d6', end_color='ffd6d6', fill_type='solid')
lightBlack_fill = PatternFill(start_color='2c2932', end_color='2c2932', fill_type='solid')

#========================================================
# Configuration
#========================================================

# Column Settings for sending files.
columnNumberSendingDate = 0
columnNumberDeliveryCode_sending = 2
columnNumberExpectedCOD = 9

# Column Settings for receiving files.
columnNumberDeliveryCode_receiving = 2
columnNumberActualCOD = 6

# Addition Infomation : up to each project.

listOfAdditionalColumnName = ["customerName", "phoneNumber"]

additionalInfoColumnNumberDict = {
    listOfAdditionalColumnName[0] : 3,
    listOfAdditionalColumnName[1] : 4
}
additionalInfoValueDict = {
    listOfAdditionalColumnName[0] : "default-value",
    listOfAdditionalColumnName[1] : "default-value"
}

sendingColumnBias = -1 #from summary cash (excel structure - footer).

receivingHeaderBias = 12 #for delivery header
receivingFooterBias = -14 #for delivery footer

#--------------------------------------------------------
# Implementation
#--------------------------------------------------------
logInfo("intro")
sendings = getSendingInfo(excelProductSendingPath, sendingColumnBias)
receivingDict = getReceivingInfo()
updateReport()
sendingDatabase = getSendingInfo(reportFolder, 0) #re-use this function, so have some wired structure.
paidList, nonPaidList, remainingReceivingDict = getPaymentList(receivingDict)
trackCOD()
moveUsedFiles()
logInfo("outro")
